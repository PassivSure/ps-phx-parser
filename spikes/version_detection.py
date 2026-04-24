"""Prototype: detect a PHPP workbook's version + language from its Data sheet.

Ports PHX.PHPP.phpp_app.PHPPConnection.get_phpp_version to openpyxl so it
works headless. Not production code — see `spikes/version-detection.md`.

Usage: uv run python spikes/version_detection.py <phpp.xlsx>
"""

from __future__ import annotations

import pathlib
import sys
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DATA_SHEET_NAMES = {"DATA", "DATEN", "DATOS"}
LANGUAGE_PROXIES = {
    "1-PE-FAKTOREN": "DE",
    "1-FACTORES EP": "ES",
    "1-PE-FACTORS": "EN",
}


class DetectionError(Exception):
    pass


@dataclass
class DetectedVersion:
    major: str
    minor: str
    language: str

    @property
    def shape_stem(self) -> str:
        """PHX's `{LANG}_{major}_{minor}` with clean_input normalization."""
        major = self.major.upper().strip().replace(" ", "").replace(".", "_")
        minor = self.minor.upper().strip().replace(" ", "").replace(".", "_")
        return f"{self.language}_{major}_{minor}"


def _find_data_sheet(wb) -> Worksheet:
    sheets_upper = {s.upper(): s for s in wb.sheetnames}
    for candidate in DATA_SHEET_NAMES:
        if candidate in sheets_upper:
            return wb[sheets_upper[candidate]]
    raise DetectionError(f"No Data/Daten/Datos sheet. Sheets: {wb.sheetnames[:8]}")


def _find_phpp_row(ws: Worksheet, col: int = 1, row_range: range = range(1, 11)) -> int:
    for r in row_range:
        v = ws.cell(row=r, column=col).value
        if v and str(v).upper().strip().replace(" ", "").startswith("PHPP"):
            return r
    raise DetectionError(
        f"No 'PHPP' marker in col A rows {row_range.start}-{row_range.stop - 1} "
        f"of sheet {ws.title!r}. Likely pre-v9 PHPP — not supported."
    )


def detect_version(path: pathlib.Path) -> DetectedVersion:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = _find_data_sheet(wb)
        row = _find_phpp_row(ws)
        row_vals = [ws.cell(row=row, column=c).value for c in range(1, 20)]
        non_blank = [v for v in row_vals if v not in (None, "")]

        if len(non_blank) < 2:
            raise DetectionError(f"PHPP row {row} has no version cell: {non_blank}")

        raw_version = str(non_blank[1])
        if "." not in raw_version:
            raise DetectionError(f"Unexpected version cell {raw_version!r}")
        major, minor = raw_version.split(".", 1)

        last = str(non_blank[-1]).upper().strip()
        language = None
        for needle, lang in LANGUAGE_PROXIES.items():
            if needle in last:
                language = lang
                break
        if language is None:
            # Pre-v10 files lack the PE-factor proxy. Default to EN since
            # we only ship EN shape files anyway. See spike writeup.
            language = "EN"

        return DetectedVersion(major=major, minor=minor, language=language)
    finally:
        wb.close()


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: version_detection.py <phpp.xlsx>", file=sys.stderr)
        return 2
    path = pathlib.Path(sys.argv[1]).expanduser()
    if not path.exists():
        print(f"missing: {path}", file=sys.stderr)
        return 1
    try:
        v = detect_version(path)
    except DetectionError as e:
        print(f"DETECTION FAILED: {e}", file=sys.stderr)
        return 1
    print(f"raw:   major={v.major!r} minor={v.minor!r} language={v.language!r}")
    print(f"shape: {v.shape_stem}.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
