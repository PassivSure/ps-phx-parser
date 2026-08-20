"""Shared test fixtures."""

import time
from io import BytesIO

import pytest
from openpyxl import Workbook
from PHX.PHPP.phpp_localization import shape_model

from app.parser import load_shape

# Synthetic ground-truth values for the KPI fixture. Real-file ground truth
# lives in tests/test_kpis_real_workbook.py and is sampled from
# ~/Downloads/PHPP_V10.6_IP_Example.xlsx.
KPI_FIXTURE_VALUES = {
    "tfa": 2400.0,
    "heating_demand_specific": 4.2,
    "cooling_demand_specific": 2.1,
    "heating_peak_w1": 22000.0,
    "heating_peak_w2": 24000.0,  # max should win
    "cooling_peak_w1": 18000.0,
    "cooling_peak_w2": 18000.0,
    "per_total_v": 7.3,  # PER source EUI
    "per_total_x": 15.2,  # PE demand
}


@pytest.fixture
def shape_en_10_6ip() -> shape_model.PhppShape:
    return load_shape("EN_10_6IP")


def build_workbook_bytes(
    shape: shape_model.PhppShape,
    num_of_units_value: int | float | str | None = 42,
    data_sheet_version: str | None = None,
    data_sheet_pe_factor: str | None = "1-PE-factors (non-renewable) PHI Certification",
    with_kpis: bool = False,
    cooling_peak_value: float | None = None,
    with_envelope: bool = False,
    with_project_info: bool = False,
) -> bytes:
    """Build a minimal in-memory PHPP-shaped workbook for tests.

    Always writes:
      - A sheet named after shape.VERIFICATION.name
      - shape.VERIFICATION.num_of_units locator string + value cell

    Optionally writes a Data sheet for version-detection tests. Pass
    data_sheet_version='10.6 IP' (or similar) to enable it.

    Optionally writes the four KPI sheets (Heating, Cooling, Heating load,
    Cooling load, PER) with values from KPI_FIXTURE_VALUES — pass
    with_kpis=True. Pass cooling_peak_value to override the cooling peak
    (e.g. -50 to test the negative-sign gotcha).
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(shape.VERIFICATION.name)
    item = shape.VERIFICATION.num_of_units
    locator_row = 30
    ws[f"{item.locator_col}{locator_row}"] = item.locator_string
    value_row = locator_row + item.input_row_offset
    if num_of_units_value is not None:
        ws[f"{item.input_column}{value_row}"] = num_of_units_value

    if data_sheet_version is not None:
        data_ws = wb.create_sheet("Data")
        data_ws["A5"] = "PHPP Version"
        data_ws["B5"] = data_sheet_version
        if data_sheet_pe_factor:
            data_ws["C5"] = "Language"
            data_ws["D5"] = data_sheet_pe_factor

    if with_kpis:
        # Heating sheet: annual demand at col_kWh_m2_year x row_annual_demand
        hd = shape.HEATING_DEMAND
        h_ws = wb.create_sheet(hd.name)
        h_ws[f"{hd.col_kWh_m2_year}{hd.row_annual_demand}"] = (
            KPI_FIXTURE_VALUES["heating_demand_specific"]
        )

        # Cooling sheet — TFA at address_tfa, sensible demand
        cd = shape.COOLING_DEMAND
        c_ws = wb.create_sheet(cd.name)
        c_ws[cd.address_tfa] = KPI_FIXTURE_VALUES["tfa"]
        c_ws[f"{cd.col_kWh_m2_year}{cd.row_annual_sensible_demand}"] = (
            KPI_FIXTURE_VALUES["cooling_demand_specific"]
        )

        # Heating load sheet — w1 + w2 at row_total_load
        hp = shape.HEATING_PEAK_LOAD
        hp_ws = wb.create_sheet(hp.name)
        hp_ws[f"{hp.col_weather_1}{hp.row_total_load}"] = KPI_FIXTURE_VALUES["heating_peak_w1"]
        hp_ws[f"{hp.col_weather_2}{hp.row_total_load}"] = KPI_FIXTURE_VALUES["heating_peak_w2"]

        # Cooling load sheet
        cp = shape.COOLING_PEAK_LOAD
        cp_ws = wb.create_sheet(cp.name)
        cp_w1 = (
            cooling_peak_value
            if cooling_peak_value is not None
            else KPI_FIXTURE_VALUES["cooling_peak_w1"]
        )
        cp_w2 = (
            cooling_peak_value
            if cooling_peak_value is not None
            else KPI_FIXTURE_VALUES["cooling_peak_w2"]
        )
        cp_ws[f"{cp.col_weather_1}{cp.row_total_sensible_load}"] = cp_w1
        cp_ws[f"{cp.col_weather_2}{cp.row_total_sensible_load}"] = cp_w2

        # PER sheet — locator-string-driven row scan looks for "Total energy demand"
        per = shape.PER
        per_ws = wb.create_sheet(per.name)
        total_row = 68  # arbitrary; the parser finds it by string
        per_ws[f"{per.locator_col}{total_row}"] = "Total energy demand kBTU/(ft²yr)"
        per_ws[f"{per.columns.per_energy}{total_row}"] = KPI_FIXTURE_VALUES["per_total_v"]
        per_ws[f"{per.columns.pe_energy}{total_row}"] = KPI_FIXTURE_VALUES["per_total_x"]

    if with_envelope:
        # Areas sheet: surface_rows + thermal_bridge_rows + airtightness via
        # the same locator pattern the real workbook uses.
        areas = shape.AREAS
        a_ws = wb.create_sheet(areas.name)

        # Surface header and entries
        sr = areas.surface_rows
        sr_header_col = sr.locator_col_header  # K
        sr_header_row = 31
        a_ws[f"{sr_header_col}{sr_header_row}"] = sr.locator_string_header
        # Skip header label row at +1; real entries start at +2
        sr_entries = [
            # (group_raw, description, area)
            ("8-External wall - ambient", "Wall_North", 200.0),
            ("10-Roof / ceiling - ambient", "Roof_Main", 1500.0),
            ("11-Floor slab / basement ceiling", "Slab_Ground", 1200.0),
            ("7-Exterior door", "Front_Door", 21.0),
            ("0-Projected building footprint", "Footprint_Aggregate", 9999.0),  # skipped
            ("2-Fenster Nord", "Window_Placeholder", None),  # skipped
        ]
        for i, (group, desc, area) in enumerate(sr_entries):
            r = sr_header_row + 2 + i
            a_ws[f"{sr.inputs.group_number.column}{r}"] = group
            a_ws[f"{sr.inputs.description.column}{r}"] = desc
            if area is not None:
                a_ws[f"{sr.inputs.area.column}{r}"] = area

        # Thermal bridges
        tb = areas.thermal_bridge_rows
        tb_header_row = sr_header_row + 2 + len(sr_entries) + 5  # gap
        a_ws[f"{tb.locator_col_header}{tb_header_row}"] = tb.locator_string_header
        tb_entries = [("Slab edge", 110.0, 0.05), ("Window jamb", 65.0, 0.02)]
        for i, (desc, length, psi) in enumerate(tb_entries):
            r = tb_header_row + 2 + i
            a_ws[f"{tb.inputs.description.column}{r}"] = desc
            a_ws[f"{tb.inputs.length.column}{r}"] = length
            a_ws[f"{tb.inputs.psi_value.column}{r}"] = psi

        # Ventilation sheet — n50 via locator pattern
        vent = shape.VENTILATION
        v_ws = wb.create_sheet(vent.name)
        n50_row = 23
        v_ws[f"{vent.airtightness_n50.locator_col}{n50_row}"] = (
            vent.airtightness_n50.locator_string
        )
        v_ws[f"{vent.airtightness_n50.input_column}{n50_row}"] = 0.6

    if with_project_info:
        # Overview sheet: project name, organizations
        ov = shape.OVERVIEW
        ov_ws = wb.create_sheet(ov.name)
        ov_ws[ov.basic_data.address_project_name] = "Test Passive House"
        # Organizations live at pinned rows in Overview cols B + C
        org_rows = {
            31: ("Home owner name / E-mail", "Acme Owners LLC"),
            33: ("Architect name / E-mail", "Smith Architects"),
            34: ("Mechanical engineer name / E-mail", "Cool MEP Inc"),
            35: ("Energy consultant name / E-mail", "EnergyPro"),
            48: ("Certification body Name / E-mail", "PHIUS"),
        }
        for row, (label, name) in org_rows.items():
            ov_ws[f"B{row}"] = label
            ov_ws[f"C{row}"] = name

        # Verification sheet: address fields + occupancy via PHX shape
        vf_ws = wb["Verification"] if "Verification" in wb.sheetnames else wb.create_sheet(
            shape.VERIFICATION.name
        )
        vf_ws["K6"] = "100 Main St"
        vf_ws["K7"] = '"94028"'  # PHPP wraps zips in literal quotes
        vf_ws["L7"] = "Sunnyvale"
        vf_ws["K8"] = "CA"

        # phi_building_use_type via locator pattern
        use_type = shape.VERIFICATION.phi_building_use_type
        # Pick a row not already used by num_of_units (which uses row 30)
        use_locator_row = 50
        vf_ws[f"{use_type.locator_col}{use_locator_row}"] = use_type.locator_string
        vf_ws[
            f"{use_type.input_column}{use_locator_row + use_type.input_row_offset}"
        ] = "10-Residential building: Residential"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def workbook_bytes(shape_en_10_6ip) -> bytes:
    return build_workbook_bytes(shape_en_10_6ip, num_of_units_value=42)


@pytest.fixture
def workbook_bytes_with_version(shape_en_10_6ip) -> bytes:
    return build_workbook_bytes(
        shape_en_10_6ip,
        num_of_units_value=42,
        data_sheet_version="10.6 IP",
    )


@pytest.fixture
def workbook_bytes_with_kpis(shape_en_10_6ip) -> bytes:
    return build_workbook_bytes(shape_en_10_6ip, with_kpis=True)


@pytest.fixture
def workbook_bytes_with_envelope(shape_en_10_6ip) -> bytes:
    return build_workbook_bytes(shape_en_10_6ip, with_envelope=True)


@pytest.fixture
def workbook_bytes_with_project_info(shape_en_10_6ip) -> bytes:
    return build_workbook_bytes(shape_en_10_6ip, with_project_info=True)


# ── Async /parse helpers ────────────────────────────────────────────────────
#
# /parse returns 202 + a job id and the work continues on a parse thread, so any
# test wanting a parsed body has to poll — exactly as the real ps-rails client
# does. Bounded so a hang fails the suite rather than hanging it.
POLL_ATTEMPTS = 100
POLL_INTERVAL_SECONDS = 0.05


@pytest.fixture(autouse=True)
def _clean_job_store():
    """Job state is process-global, so leaking it between examples would let one
    test see another's jobs."""
    from app import jobs

    jobs._reset_for_tests()
    yield
    jobs._reset_for_tests()


def start_parse(client, url, **payload) -> str:
    """POST /parse and return the job id, asserting the 202 accept contract."""
    response = client.post("/parse", json={"url": url, **payload})
    assert response.status_code == 202, response.text
    body = response.json()
    assert body["status"] == "pending"
    return body["job_id"]


def poll_until_settled(client, job_id: str) -> dict:
    """Poll /parse/{job_id} until it leaves `pending`; return the final body."""
    for _ in range(POLL_ATTEMPTS):
        response = client.get(f"/parse/{job_id}")
        assert response.status_code == 200
        body = response.json()
        if body["status"] != "pending":
            return body
        time.sleep(POLL_INTERVAL_SECONDS)
    raise AssertionError(f"job {job_id} never settled")


def parse_and_wait(client, url, **payload) -> dict:
    """Start a parse and block until it settles. Returns the settled body."""
    return poll_until_settled(client, start_parse(client, url, **payload))
