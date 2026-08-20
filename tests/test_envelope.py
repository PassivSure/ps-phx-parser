"""Unit tests for app.envelope on synthetic workbooks."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.envelope import (
    _classify,
    _computed_area_col,
    _group_number,
    _surface_components,
    read_envelope,
)
from app.parser import load_shape


def _envelope(workbook_bytes: bytes) -> dict:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    try:
        return read_envelope(wb, load_shape("EN_10_6IP"))
    finally:
        wb.close()


def test_envelope_returns_components_and_airtightness(workbook_bytes_with_envelope):
    env = _envelope(workbook_bytes_with_envelope)
    assert "components" in env
    assert "airtightness" in env


def test_surface_classification_from_group_prefix(workbook_bytes_with_envelope):
    """Group integer prefix in column M maps to schema component enum.
    Aggregate groups (0=footprint, 1=TFA) and window placeholders (2-6)
    are skipped — windows handled separately in P2.3.c."""
    env = _envelope(workbook_bytes_with_envelope)
    # Filter to just the opaque surfaces (not thermal bridges)
    opaque = [c for c in env["components"] if c["component"] != "thermal_bridge"]
    by_label = {c["label"]: c for c in opaque}

    assert by_label["Wall_North"]["component"] == "wall"
    assert by_label["Roof_Main"]["component"] == "roof"
    assert by_label["Slab_Ground"]["component"] == "slab_on_grade"
    assert by_label["Front_Door"]["component"] == "door"

    # Skipped placeholders: footprint and window placeholders
    assert "Footprint_Aggregate" not in by_label
    assert "Window_Placeholder" not in by_label


def test_surface_areas_emitted(workbook_bytes_with_envelope):
    env = _envelope(workbook_bytes_with_envelope)
    by_label = {c["label"]: c for c in env["components"]}
    assert by_label["Wall_North"]["area_ft2"] == 200.0
    assert by_label["Roof_Main"]["area_ft2"] == 1500.0
    assert by_label["Slab_Ground"]["area_ft2"] == 1200.0


def test_thermal_bridges_emit_length_and_psi(workbook_bytes_with_envelope):
    env = _envelope(workbook_bytes_with_envelope)
    tbs = [c for c in env["components"] if c["component"] == "thermal_bridge"]
    assert len(tbs) == 2

    by_label = {tb["label"]: tb for tb in tbs}
    assert by_label["Slab edge"]["length_ft"] == 110.0
    assert by_label["Slab edge"]["psi_value_Btuh_ftF"] == 0.05
    assert by_label["Window jamb"]["length_ft"] == 65.0


def test_airtightness_n50_read_from_ventilation(workbook_bytes_with_envelope):
    env = _envelope(workbook_bytes_with_envelope)
    assert env["airtightness"]["n50_ach"] == 0.6
    assert env["airtightness"]["source"] == "Ventilation!M23"


def test_envelope_handles_missing_sheets(workbook_bytes):
    """workbook_bytes has no Areas/Ventilation sheets — emit empty
    components list and null airtightness."""
    env = _envelope(workbook_bytes)
    assert env["components"] == []
    assert env["airtightness"]["n50_ach"] is None


def test_iterate_until_blank_terminates_on_blank_description(shape_en_10_6ip):
    """A blank description column terminates the surface walk — no
    spurious entries from rows past the user's last input."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    areas = shape_en_10_6ip.AREAS
    ws = wb.create_sheet(areas.name)

    sr = areas.surface_rows
    header_row = 30
    ws[f"{sr.locator_col_header}{header_row}"] = sr.locator_string_header

    # 3 entries, then blank, then a stray entry that must NOT be included
    entries = [
        (header_row + 2, "8-External wall - ambient", "Wall_A", 100.0),
        (header_row + 3, "8-External wall - ambient", "Wall_B", 200.0),
        (header_row + 4, "10-Roof / ceiling - ambient", "Roof_A", 1500.0),
        # Row +5 deliberately blank (terminator)
        (header_row + 6, "8-External wall - ambient", "STRAY_AFTER_BLANK", 999.0),
    ]
    for r, group, desc, area in entries:
        ws[f"{sr.inputs.group_number.column}{r}"] = group
        ws[f"{sr.inputs.description.column}{r}"] = desc
        ws[f"{sr.inputs.area.column}{r}"] = area

    buf = BytesIO()
    wb.save(buf)

    env = _envelope(buf.getvalue())
    labels = [c["label"] for c in env["components"]]
    assert "Wall_A" in labels
    assert "Roof_A" in labels
    assert "STRAY_AFTER_BLANK" not in labels


# --- group-number parsing --------------------------------------------------
#
# v10 writes the labelled dropdown member as text; v9.7 writes the bare number.
# Accepting only the text form skipped every surface on every v9.7 file.


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("8-External wall - ambient", 8),  # v10
        ("11-Floor slab / Basement ceiling", 11),
        (8, 8),  # v9.7 — openpyxl hands back an int
        (11, 11),
        (8.0, 8),  # defensively: a float that is a whole number
        ("8", 8),  # a numeric string
        ("  8  ", 8),
        (8.5, None),  # not a group number
        (None, None),
        ("", None),
        ("wall", None),
        (True, None),  # bool is an int subclass; a checkbox is not a group
    ],
)
def test_group_number_accepts_both_forms(raw, expected):
    assert _group_number(raw) == expected


@pytest.mark.parametrize("raw", [0, 1, 2, 6, "0-Projected building footprint"])
def test_aggregate_and_window_groups_are_still_skipped(raw):
    """The skip must not depend on which form the cell takes — a bare 0 is as
    much an aggregate row as '0-Projected building footprint'."""
    assert _classify(raw) is None


@pytest.mark.parametrize(
    "raw,expected",
    [(7, "door"), (8, "wall"), (9, "wall"), (10, "roof"), (11, "slab_on_grade")],
)
def test_classify_maps_bare_numbers(raw, expected):
    assert _classify(raw) == expected


# --- computed area column --------------------------------------------------


def _areas_sheet(area_header="Area [ft²]"):
    """Surface table with BOTH an override column and PHPP's computed one, so
    a test can tell which was read."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Areas"
    ws["K31"] = "Area input"
    ws["L32"] = "Building assembly description"
    ws["M32"] = "To group No."
    ws["T32"] = "User-defined calculation [ft²]"
    if area_header:
        ws["Z32"] = area_header
    ws["L33"] = "Wall_001"
    ws["M33"] = 8  # bare int, the v9.7 form
    ws["T33"] = 111.0  # override — empty in real files for computed surfaces
    ws["Z33"] = 999.0  # PHPP's own result
    return wb


def test_computed_area_column_is_located_by_label():
    wb = _areas_sheet()
    assert _computed_area_col(wb["Areas"], 32) == 26  # Z


def test_computed_area_column_absent_returns_none():
    wb = _areas_sheet(area_header=None)
    assert _computed_area_col(wb["Areas"], 32) is None


def test_surfaces_prefer_phpps_computed_area_over_the_override(shape_en_10_6ip):
    """The shape's `area` input points at the USER OVERRIDE, which is empty for
    every surface PHPP computes itself — most of them. Reading it emitted
    area_ft2=None for 40 of 41 surfaces on a real v10 file."""
    wb = _areas_sheet()
    surfaces = _surface_components(wb, shape_en_10_6ip.AREAS)

    assert len(surfaces) == 1
    assert surfaces[0]["area_ft2"] == 999.0
    assert surfaces[0]["source_area"] == "Areas!Z33"


def test_falls_back_to_the_override_when_there_is_no_computed_column(
    shape_en_10_6ip,
):
    wb = _areas_sheet(area_header=None)
    surfaces = _surface_components(wb, shape_en_10_6ip.AREAS)

    assert surfaces[0]["area_ft2"] == 111.0
    assert surfaces[0]["source_area"] == "Areas!T33"
