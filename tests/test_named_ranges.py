from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from app.named_ranges import resolve


def _wb_with_names():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventilation SI"
    ws["K91"] = "01ud-Swegon Casa R7"
    ws["L91"] = None
    ws["M91"] = "extra"
    ws["B4"] = 306.924
    wb.defined_names.add(DefinedName("single", attr_text="'Ventilation SI'!$B$4"))
    wb.defined_names.add(DefinedName("multi", attr_text="'Ventilation SI'!$K$91:$M$91"))
    wb.defined_names.add(DefinedName("bad_sheet", attr_text="'No Such Sheet'!$A$1"))
    wb.defined_names.add(DefinedName("unparseable", attr_text="#REF!"))
    return wb


def test_single_cell_returns_the_scalar():
    assert resolve(_wb_with_names(), "single") == 306.924


def test_multi_cell_returns_all_non_empty_values_not_just_the_first():
    # The defect this guards: taking the first cell is right often enough
    # to look correct, and wrong wherever the useful value is not first.
    assert resolve(_wb_with_names(), "multi") == ["01ud-Swegon Casa R7", "extra"]


def test_absent_name_returns_none():
    assert resolve(_wb_with_names(), "nope") is None


def test_missing_sheet_returns_none_rather_than_raising():
    assert resolve(_wb_with_names(), "bad_sheet") is None


def test_unparseable_reference_returns_none_rather_than_raising():
    assert resolve(_wb_with_names(), "unparseable") is None


def test_multi_cell_that_is_entirely_empty_returns_none():
    wb = _wb_with_names()
    wb["Ventilation SI"]["K91"] = None
    wb["Ventilation SI"]["M91"] = None
    assert resolve(wb, "multi") is None


def test_range_with_single_populated_cell_returns_list_not_scalar():
    # The return shape is keyed on reference syntax (presence of `:` in the reference),
    # not on how many cells contain data. A range reference with one populated cell
    # must return a one-element list, not the scalar.
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws["B4"] = 42.0
    ws["B5"] = None
    wb.defined_names.add(DefinedName("single_populated", attr_text="'Test'!$B$4:$B$5"))
    assert resolve(wb, "single_populated") == [42.0]


def test_degenerate_range_reference_returns_list():
    # A degenerate range reference like $B$4:$B$4 (same cell on both sides)
    # is still a range syntactically and must return a list, not the scalar.
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws["B4"] = 100.5
    wb.defined_names.add(DefinedName("degenerate", attr_text="'Test'!$B$4:$B$4"))
    assert resolve(wb, "degenerate") == [100.5]
