"""Envelope areas reconciled against PHPP's own group subtotals.

Skipped when the files aren't present. This is the strongest check available
for the surface reader: PHPP publishes a Summary block on the Areas sheet
giving total area per group, computed by the workbook itself. Summing what we
parse per component type and comparing against that block tests the reads
against the spreadsheet's own arithmetic rather than against numbers we once
wrote down.

It covers both fixes that made these numbers real:

  * v9.7 writes bare integers in the group column where v10 writes labelled
    members, so every v9.7 surface used to be skipped and the components list
    came back holding thermal bridges alone.
  * The shape's ``area`` input is the user-override column, empty for any
    surface PHPP computes from its own a x b entry. On a real v10 file that
    was 40 of 41 surfaces reading ``None``.

Both would still "work" without this test — one returns an empty list, the
other returns nulls, and neither raises.
"""

import pathlib
from collections import defaultdict

import pytest

from app.envelope import GROUP_TO_COMPONENT, _find_header_row
from app.parser import parse_workbook

DOWNLOADS = pathlib.Path.home() / "Downloads"

# The Summary block's columns are transposed between the two versions:
# v9.7 has label/group/area in L/M/N, v10 has area/label/group in L/M/N.
SUMMARY_LAYOUT = {
    "EN_9_7IP": ("M", "N"),  # (group col, area col)
    "EN_10_6IP": ("N", "L"),
    "EN_10_6": ("N", "L"),
}

CASES = {
    "55th_v97": ("250121_55th_IP9_741zhOb_FINAL.xlsx", "EN_9_7IP"),
    "17mile_v97": ("250901.2821 17 mile drive.IP9.7.final.xlsx", "EN_9_7IP"),
    "6840_v106": ("260309 PHPP 6840 E ^th Ave Pkwy v 10.6.xlsx", "EN_10_6IP"),
    # easyPH, and the only file in the corpus with an unused area group --
    # its exterior-door placeholder carries no quantity, so PHPP totals the
    # group as 0 and the row must not surface as a door.
    "holmes_easyph": ("2536 Holmes_easyPH_260512.xlsx", "EN_10_6"),
}


def _phpp_group_totals(path, version):
    """Per-component totals as PHPP itself reports them, from the Summary
    block above the surface-input table."""
    from openpyxl import load_workbook

    group_col, area_col = SUMMARY_LAYOUT[version]
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb["Areas"]
        header_row = _find_header_row(ws, "K", "Area input")
        assert header_row, "surface table header not found"

        totals = defaultdict(float)
        for row in range(1, header_row):
            raw_group = ws[f"{group_col}{row}"].value
            area = ws[f"{area_col}{row}"].value
            try:
                group = int(str(raw_group).strip())
            except (TypeError, ValueError):
                continue
            if group in GROUP_TO_COMPONENT and isinstance(area, (int, float)):
                totals[GROUP_TO_COMPONENT[group]] += area
        return dict(totals)
    finally:
        wb.close()


@pytest.fixture(scope="module", params=sorted(CASES), ids=sorted(CASES))
def case(request):
    filename, version = CASES[request.param]
    path = DOWNLOADS / filename
    if not path.is_file():
        pytest.skip(f"fixture not found at {path}")
    envelope = parse_workbook(path.read_bytes(), version)["envelope"]
    return envelope, _phpp_group_totals(path, version)


def _parsed_totals(envelope):
    totals = defaultdict(float)
    for component in envelope["components"]:
        if component["component"] == "thermal_bridge":
            continue
        totals[component["component"]] += component["area_ft2"] or 0.0
    return dict(totals)


def test_every_component_total_matches_phpps_own(case):
    """Compared over the union with a 0.0 default, not by set equality: a
    component with no rows totals zero, and PHPP reports it that way. Holmes
    has no exterior door, so PHPP says `door: 0.0` while the parser emits no
    door row at all — those agree, and set equality would call them different.
    """
    envelope, phpp = case
    parsed = _parsed_totals(envelope)

    for component in set(parsed) | set(phpp):
        assert parsed.get(component, 0.0) == pytest.approx(
            phpp.get(component, 0.0), rel=1e-9, abs=1e-9
        ), component


def test_surfaces_are_present_and_all_carry_an_area(case):
    """Guards both failure modes at once: an empty list, and a full list of
    nulls. Each is silent on its own."""
    envelope, _ = case
    surfaces = [
        c for c in envelope["components"] if c["component"] != "thermal_bridge"
    ]

    assert surfaces, "no surface components parsed"
    assert all(s["area_ft2"] is not None for s in surfaces)


def test_walls_are_the_largest_opaque_group(case):
    """A weak sanity check that survives re-pinning: if group classification
    silently shifted, walls would stop dominating."""
    _, phpp = case
    assert phpp["wall"] == max(phpp.values())


def test_no_surface_is_emitted_without_an_area(case):
    """PHPP seeds one placeholder row per area group whose formula yields ""
    when the group is unused. Group 7 maps to `door`, so before this was
    handled a building with no exterior door reported one anyway, with a null
    area."""
    envelope, _ = case
    surfaces = [
        c for c in envelope["components"] if c["component"] != "thermal_bridge"
    ]
    assert all(s["area_ft2"] is not None for s in surfaces)
