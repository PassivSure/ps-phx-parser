import pytest
from openpyxl import Workbook

from app.parser import (
    ParseError,
    available_versions,
    load_shape,
    parse_workbook,
)
from tests.conftest import build_workbook_bytes


def test_available_versions_includes_en_10_6ip():
    versions = available_versions()
    assert "EN_10_6IP" in versions
    assert "EN_10_6" in versions


def test_load_shape_caches_result():
    a = load_shape("EN_10_6IP")
    b = load_shape("EN_10_6IP")
    assert a is b  # lru_cache hit


def test_load_shape_unknown_version_raises():
    with pytest.raises(ParseError, match="Unknown phpp_version"):
        load_shape("XX_99_9")


def test_parse_workbook_extracts_num_of_units(workbook_bytes):
    result = parse_workbook(workbook_bytes, "EN_10_6IP")
    assert result == {"phpp_version": "EN_10_6IP", "num_of_units": 42}


def test_parse_workbook_returns_null_when_locator_missing(shape_en_10_6ip):
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(shape_en_10_6ip.VERIFICATION.name)
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)

    result = parse_workbook(buf.getvalue(), "EN_10_6IP")
    assert result["num_of_units"] is None


def test_parse_workbook_raises_when_verification_sheet_missing(shape_en_10_6ip):
    wb = Workbook()
    wb.active.title = "RandomSheet"
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ParseError, match="Sheet 'Verification' not found"):
        parse_workbook(buf.getvalue(), "EN_10_6IP")


def test_parse_workbook_rejects_unknown_version(workbook_bytes):
    with pytest.raises(ParseError, match="Unknown phpp_version"):
        parse_workbook(workbook_bytes, "XX_99_9")


def test_parse_workbook_preserves_non_integer_values(shape_en_10_6ip):
    wb_bytes = build_workbook_bytes(shape_en_10_6ip, num_of_units_value=1.5)
    result = parse_workbook(wb_bytes, "EN_10_6IP")
    assert result["num_of_units"] == 1.5
