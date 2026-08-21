import json
import pathlib

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from app.equipment import _classify_ventilation, _strip_prefix, read_equipment

SCHEMA_PATH = pathlib.Path(__file__).resolve().parent.parent / "schema" / "output.schema.json"


def _hvac_equipment_item_allowed_keys() -> set[str]:
    """The live contract, not a transcription of it. Reading
    $defs/hvac_equipment_item/properties out of schema/output.schema.json
    directly means this test checks against the real schema and can't
    silently drift from a hardcoded copy of its key set."""
    schema = json.loads(SCHEMA_PATH.read_text())
    return set(schema["$defs"]["hvac_equipment_item"]["properties"])


def _wb(names: dict[str, object], sheet: str = "Ventilation SI"):
    """Build a workbook whose defined names point at real cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i, (name, value) in enumerate(names.items(), start=2):
        ws.cell(row=i, column=1, value=value)
        wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!$A${i}"))
    return wb


class TestStripPrefix:
    # PHPP dropdown members carry an id prefix. The same prefix format already
    # bit the easyPH writer, so it is parsed rather than assumed away.
    @pytest.mark.parametrize(
        "raw,expected",
        [
            ("01ud-Swegon Casa R7 Genius Sorption", "Swegon Casa R7 Genius Sorption"),
            ("1363vs03-Brink Climate Systems B.V. - Brink Flair",
             "Brink Climate Systems B.V. - Brink Flair"),
            ("99ud-Standard air-to-air heat pump", "Standard air-to-air heat pump"),
            ("no prefix here", "no prefix here"),
            ("", None),
            (None, None),
        ],
    )
    def test_strips_only_the_leading_id(self, raw, expected):
        assert _strip_prefix(raw) == expected

    def test_keeps_internal_hyphens(self):
        # "Brink Climate Systems B.V. - Brink Flair" contains a hyphen that is
        # part of the name; splitting on the last one would truncate it.
        assert _strip_prefix("1363vs03-A - B") == "A - B"


class TestVentilationClassification:
    # Humidity recovery is the hrv/erv discriminator. Both units in the real
    # corpus recover moisture, so the hrv branch needs a synthetic case or it
    # is never exercised.
    def test_moisture_recovery_means_erv(self):
        assert _classify_ventilation(86.0) == "erv"

    def test_no_moisture_recovery_means_hrv(self):
        assert _classify_ventilation(0.0) == "hrv"

    def test_unknown_moisture_recovery_means_hrv(self):
        assert _classify_ventilation(None) == "hrv"


class TestDiscovery:
    def test_emits_nothing_for_a_non_v106_version(self):
        wb = _wb({"Lueftung_Auswahl_Lueftungsgeraet": "01ud-Unit"})
        assert read_equipment(wb, "EN_9_7IP") == []

    def test_emits_no_ventilation_device_when_none_is_selected(self):
        # The phantom-device guard: a file has a row per family whether or not
        # the building uses one.
        assert read_equipment(_wb({}), "EN_10_6") == []

    def test_emits_a_ventilation_device_when_one_is_selected(self):
        wb = _wb({"Lueftung_Auswahl_Lueftungsgeraet": "01ud-Swegon Casa R7"})
        items = read_equipment(wb, "EN_10_6")
        assert [i["name"] for i in items] == ["Swegon Casa R7"]

    def test_cooling_device_requires_its_presence_flag(self):
        selected = {"Kuehlgeraete_Kompressor_Umluft_Geraet": "01ud-Daikin 4MXTH36AVJU9"}
        assert read_equipment(_wb(selected), "EN_10_6") == []

        flagged = dict(selected, Kuehlgeraete_Umluft_Kuehlung_Ankreuzen="x")
        items = read_equipment(_wb(flagged), "EN_10_6")
        assert [i["equipment_type"] for i in items] == ["cooling_unit"]

    def test_every_emitted_item_uses_only_schema_keys(self):
        wb = _wb({"Lueftung_Auswahl_Lueftungsgeraet": "01ud-Swegon Casa R7"})
        allowed = _hvac_equipment_item_allowed_keys()
        for item in read_equipment(wb, "EN_10_6"):
            missing = allowed - set(item)
            extra = set(item) - allowed
            assert set(item) == allowed, f"Missing: {missing}, Extra: {extra}"
            assert item["equipment_type"] is not None

    def test_manufacturer_is_never_derived(self):
        # "Swegon" is a real manufacturer token; a naive split()[0] would emit
        # it. Verify it is always None, not invented from the name.
        wb = _wb({"Lueftung_Auswahl_Lueftungsgeraet": "01ud-Swegon Casa R7"})
        items = read_equipment(wb, "EN_10_6")
        assert len(items) == 1
        assert items[0]["name"] == "Swegon Casa R7"
        assert items[0]["manufacturer"] is None


class TestVentilationSpec:
    """Column block on the Components sheet:
    LQ id · LR description · LS heat recovery · LT humidity recovery
    · LX/LY airflow range.
    """

    def _wb_with_components(self, *, row: int, device_id: str):
        wb = Workbook()
        vent = wb.active
        vent.title = "Ventilation SI"
        vent["A2"] = f"{device_id}-Swegon Casa R7"
        wb.defined_names.add(
            DefinedName("Lueftung_Auswahl_Lueftungsgeraet",
                        attr_text="'Ventilation SI'!$A$2"))
        vent["A3"] = 306.924
        wb.defined_names.add(
            DefinedName("Lueftung_Auslegungsvolumenstrom",
                        attr_text="'Ventilation SI'!$A$3"))

        comp = wb.create_sheet("Components SI")
        comp[f"LQ{row}"] = device_id
        comp[f"LS{row}"] = 0.86
        comp[f"LT{row}"] = 0.86
        return wb

    def test_finds_a_user_defined_unit_in_the_user_block(self):
        wb = self._wb_with_components(row=13, device_id="01ud")
        item = read_equipment(wb, "EN_10_6")[0]
        assert item["heat_recovery_efficiency_pct"] == pytest.approx(86.0)
        assert item["equipment_type"] == "erv"

    def test_finds_a_certified_unit_in_the_catalog_block(self):
        # The guard: searching only the catalog passes this and fails the one
        # above, which is the commoner case in real projects.
        wb = self._wb_with_components(row=164, device_id="1363vs03")
        item = read_equipment(wb, "EN_10_6")[0]
        assert item["heat_recovery_efficiency_pct"] == pytest.approx(86.0)

    def test_emits_both_airflow_units_from_one_metric_read(self):
        wb = self._wb_with_components(row=13, device_id="01ud")
        item = read_equipment(wb, "EN_10_6")[0]
        assert item["airflow_m3h"] == pytest.approx(306.924)
        assert item["airflow_cfm"] == pytest.approx(180.65, abs=0.01)

    def test_missing_spec_row_leaves_efficiency_null_without_raising(self):
        wb = self._wb_with_components(row=13, device_id="01ud")
        del wb["Components SI"]["LS13"]
        item = read_equipment(wb, "EN_10_6")[0]
        assert item["heat_recovery_efficiency_pct"] is None
        assert item["name"] == "Swegon Casa R7"

    def test_absent_components_sheet_leaves_efficiency_null(self):
        wb = self._wb_with_components(row=13, device_id="01ud")
        del wb["Components SI"]
        item = read_equipment(wb, "EN_10_6")[0]
        assert item["heat_recovery_efficiency_pct"] is None
