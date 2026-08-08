from io import BytesIO

import pytest
from openpyxl import Workbook

from app.version_detection import (
    DetectedVersion,
    DetectionError,
    detect_version,
    strip_easyph_edition_tag,
)
from tests.conftest import build_workbook_bytes


class TestShapeStem:
    def test_normalizes_10_6_metric(self):
        v = DetectedVersion(major="10", minor="6", language="EN", raw="10.6")
        assert v.shape_stem == "EN_10_6"

    def test_normalizes_10_6_ip(self):
        v = DetectedVersion(major="10", minor="6 IP", language="EN", raw="10.6 IP")
        assert v.shape_stem == "EN_10_6IP"

    def test_normalizes_v_8_5(self):
        v = DetectedVersion(major="8", minor="5", language="EN", raw="8.5")
        assert v.shape_stem == "EN_8_5"


class TestDetectVersion:
    def test_detects_10_6_ip(self, shape_en_10_6ip):
        bytes_ = build_workbook_bytes(shape_en_10_6ip, data_sheet_version="10.6 IP")

        detected = detect_version(bytes_)

        assert detected.shape_stem == "EN_10_6IP"
        assert detected.raw == "10.6 IP"
        assert detected.language == "EN"

    def test_detects_10_6_metric(self, shape_en_10_6ip):
        bytes_ = build_workbook_bytes(shape_en_10_6ip, data_sheet_version="10.6")

        detected = detect_version(bytes_)

        assert detected.shape_stem == "EN_10_6"

    def test_defaults_to_en_when_pe_factor_is_absent(self, shape_en_10_6ip):
        bytes_ = build_workbook_bytes(
            shape_en_10_6ip,
            data_sheet_version="8.5",
            data_sheet_pe_factor=None,
        )

        detected = detect_version(bytes_)

        assert detected.language == "EN"
        assert detected.shape_stem == "EN_8_5"

    def test_raises_when_no_data_sheet(self, shape_en_10_6ip):
        bytes_ = build_workbook_bytes(shape_en_10_6ip)  # no Data sheet

        with pytest.raises(DetectionError, match="No Data/Daten/Datos sheet"):
            detect_version(bytes_)

    def test_raises_when_no_phpp_marker(self):
        wb = Workbook()
        wb.active.title = "Data"
        wb.active["A1"] = "Something else"

        buf = BytesIO()
        wb.save(buf)

        with pytest.raises(DetectionError, match="pre-v9 PHPP"):
            detect_version(buf.getvalue())

    def test_accepts_localized_daten_sheet(self, shape_en_10_6ip):
        # Build a workbook with "Daten" (German) instead of "Data"
        wb = Workbook()
        wb.remove(wb.active)
        daten = wb.create_sheet("Daten")
        daten["A5"] = "PHPP Version"
        daten["B5"] = "10.6"
        daten["D5"] = "1-PE-FAKTOREN"

        buf = BytesIO()
        wb.save(buf)

        detected = detect_version(buf.getvalue())
        assert detected.language == "DE"
        assert detected.shape_stem == "DE_10_6"

    def test_raises_on_unexpected_version_cell(self):
        wb = Workbook()
        wb.active.title = "Data"
        wb.active["A5"] = "PHPP Version"
        wb.active["B5"] = "garbage"

        buf = BytesIO()
        wb.save(buf)

        with pytest.raises(DetectionError, match="Unexpected version cell"):
            detect_version(buf.getvalue())


class TestEasyPhEditionTag:
    """easyPH tags the edition onto the version cell on 'Data'.

    The tag names the EDITION, not a different shapefile: an easyPH workbook is
    a standard PHPP plus an 'easyPH' input worksheet, so it uses its base
    version's shape. PHX agrees — is_easyPh() detects the edition by looking for
    the worksheet, never by reading this string.

    Left in, the tag lands in `minor` and shape_stem yields
    EN_10_6EASYPHV3IP, a shapefile that does not and should not exist.
    """

    @pytest.mark.parametrize(
        "raw,expected",
        [
            ("10.6 easyPHv3", "10.6"),
            ("10.6 easyPHv3 IP", "10.6 IP"),
            ("10.6easyPHv3", "10.6"),
            ("10.6 EASYPHV3", "10.6"),
            ("10.6 easyPH", "10.6"),
            # standard version cells are untouched
            ("10.6", "10.6"),
            ("10.6 IP", "10.6 IP"),
            ("9.6a", "9.6a"),
        ],
    )
    def test_strips_only_the_edition_tag(self, raw, expected):
        assert strip_easyph_edition_tag(raw) == expected

    def test_easyph_metric_resolves_to_its_base_shape(self, shape_en_10_6ip):
        bytes_ = build_workbook_bytes(
            shape_en_10_6ip, data_sheet_version="10.6 easyPHv3"
        )

        detected = detect_version(bytes_)

        assert detected.shape_stem == "EN_10_6"
        # raw keeps the original, so a caller can still tell it was easyPH
        assert detected.raw == "10.6 easyPHv3"

    def test_easyph_ip_keeps_its_unit_suffix(self, shape_en_10_6ip):
        """The IP suffix IS part of the shape — only the edition tag goes."""
        bytes_ = build_workbook_bytes(
            shape_en_10_6ip, data_sheet_version="10.6 easyPHv3 IP"
        )

        detected = detect_version(bytes_)

        assert detected.shape_stem == "EN_10_6IP"
        assert detected.raw == "10.6 easyPHv3 IP"
