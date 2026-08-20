"""Unit tests for app.kpis on synthetic workbooks."""

from app.kpis import _find_v9_summary_row, _per_total, read_kpis
from app.parser import load_shape
from tests.conftest import KPI_FIXTURE_VALUES, build_workbook_bytes


def _kpis(workbook_bytes: bytes) -> dict:
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    try:
        return read_kpis(wb, load_shape("EN_10_6IP"))
    finally:
        wb.close()


def test_tfa_read_from_cooling_sheet(workbook_bytes_with_kpis):
    kpis = _kpis(workbook_bytes_with_kpis)
    assert kpis["tfa"]["value"] == KPI_FIXTURE_VALUES["tfa"]
    assert kpis["tfa"]["unit"] == "ft2"
    assert kpis["tfa"]["source"] == "Cooling!O8"


def test_specific_demands_read_from_demand_sheets(workbook_bytes_with_kpis):
    kpis = _kpis(workbook_bytes_with_kpis)
    assert kpis["heating_demand"]["value"] == KPI_FIXTURE_VALUES["heating_demand_specific"]
    assert kpis["heating_demand"]["unit"] == "kBtu/ft2yr"
    assert kpis["cooling_demand"]["value"] == KPI_FIXTURE_VALUES["cooling_demand_specific"]
    assert kpis["cooling_demand"]["unit"] == "kBtu/ft2yr"


def test_peak_heating_returns_max_of_two_weather_scenarios(workbook_bytes_with_kpis):
    """w1=22000, w2=24000 → design peak is the worse case (max=24000)."""
    kpis = _kpis(workbook_bytes_with_kpis)
    assert kpis["peak_loads"]["heating"]["value"] == KPI_FIXTURE_VALUES["heating_peak_w2"]
    assert kpis["peak_loads"]["heating"]["unit"] == "Btu/h"


def test_peak_cooling_negative_value_returns_null(shape_en_10_6ip):
    """Cooling-load sign gotcha: negatives are heat-loss numbers, not real
    cooling demand. Return None instead of leaking a misleading sign."""
    workbook = build_workbook_bytes(
        shape_en_10_6ip, with_kpis=True, cooling_peak_value=-500.0
    )
    kpis = _kpis(workbook)
    assert kpis["peak_loads"]["cooling"]["value"] is None


def test_per_totals_found_by_locator_string(workbook_bytes_with_kpis):
    """PER row 68 is found by `Total energy demand` prefix match in column P,
    not by hardcoded row — the table shifts when users add heating types."""
    kpis = _kpis(workbook_bytes_with_kpis)
    assert kpis["source_eui"]["value"] == KPI_FIXTURE_VALUES["per_total_v"]
    assert kpis["source_eui"]["unit"] == "kBtu/ft2yr"
    assert kpis["pe_demand"]["value"] == KPI_FIXTURE_VALUES["per_total_x"]
    assert kpis["pe_demand"]["unit"] == "kBtu/ft2yr"


def test_per_totals_locator_works_when_row_shifts(shape_en_10_6ip):
    """Author the PER total at a non-default row to prove we don't depend
    on row 68."""
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    per = shape_en_10_6ip.PER
    ws = wb.create_sheet(per.name)
    custom_row = 95
    ws[f"{per.locator_col}{custom_row}"] = "Total energy demand kBTU/(ft²yr)"
    ws[f"{per.columns.per_energy}{custom_row}"] = 11.1
    ws[f"{per.columns.pe_energy}{custom_row}"] = 22.2

    # Also need a Verification sheet so parse_workbook doesn't choke
    # downstream — read_kpis itself doesn't, but cleaner this way.
    wb.create_sheet(shape_en_10_6ip.VERIFICATION.name)

    buf = BytesIO()
    wb.save(buf)

    kpis = _kpis(buf.getvalue())
    assert kpis["source_eui"]["value"] == 11.1
    assert kpis["pe_demand"]["value"] == 22.2
    assert kpis["source_eui"]["source"] == f"PER!V{custom_row}"


class TestV9PerSummaryLocator:
    """The v9 PER summary sits ABOVE the breakdown and carries no
    "Total energy demand" label, so it is found by the header in its own
    column. See app.kpis._find_v9_summary_row.
    """

    @staticmethod
    def _v9_per_sheet(summary_value=14.5, pe_value=33.4):
        """A minimal v9-shaped PER sheet: header at 15, units 16, factor-set
        17, summary 18, breakdown from 20."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "PER"
        ws["P14"] = "Energy demand"
        ws["W15"] = "PER specific value"
        ws["Y15"] = "PE Value"
        ws["W16"] = "kBTU/(ft²yr)"
        ws["Y16"] = "kBTU/(ft²yr)"
        ws["X17"] = "1-PE-factors (non-renewable) PHI Certification"
        if summary_value is not None:
            ws["W18"] = summary_value
        if pe_value is not None:
            ws["Y18"] = pe_value
        # the per-end-use breakdown, which must never be mistaken for a total
        ws["P20"] = "Heating"
        ws["W20"] = 1.9294808282904012
        ws["Y20"] = 4.560591048686403
        return wb

    def test_finds_the_summary_by_its_in_column_header(self):
        wb = self._v9_per_sheet()
        assert _find_v9_summary_row(wb["PER"], 23, "per_energy") == 18  # W
        assert _find_v9_summary_row(wb["PER"], 25, "pe_energy") == 18  # Y

    def test_returns_none_rather_than_the_breakdown_when_the_summary_is_blank(self):
        """The guard that matters. W20 holds the heating row — a plausible
        fraction of the right answer, which is exactly what makes returning it
        worse than returning nothing."""
        wb = self._v9_per_sheet(summary_value=None)
        assert _find_v9_summary_row(wb["PER"], 23, "per_energy") is None

    def test_returns_none_when_the_header_is_absent(self):
        wb = self._v9_per_sheet()
        del wb["PER"]["W15"]
        assert _find_v9_summary_row(wb["PER"], 23, "per_energy") is None

    def test_returns_none_for_a_column_with_no_known_header(self):
        wb = self._v9_per_sheet()
        assert _find_v9_summary_row(wb["PER"], 20, "final_energy") is None

    def test_per_total_actually_uses_the_fallback(self):
        """Covers the WIRING, not just the helper.

        The real-workbook tests that prove this end to end are skipif-gated on
        local files, so without this CI would stay green if _per_total stopped
        calling the fallback at all."""
        per = load_shape("EN_9_7IP").PER
        wb = self._v9_per_sheet(summary_value=14.5, pe_value=33.4)

        assert _per_total(wb, per, "per_energy")["value"] == 14.5
        assert _per_total(wb, per, "pe_energy")["value"] == 33.4
        assert _per_total(wb, per, "per_energy")["source"] == "PER!W18"
