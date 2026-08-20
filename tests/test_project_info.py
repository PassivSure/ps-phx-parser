"""Unit tests for app.project_info on synthetic workbooks."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.parser import load_shape
from app.project_info import (
    _address_fields,
    _address_row,
    _is_organization_label,
    _organizations,
    read_project_info,
)


def _info(workbook_bytes: bytes) -> dict:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    try:
        return read_project_info(wb, load_shape("EN_10_6IP"))
    finally:
        wb.close()


def test_project_name_from_overview(workbook_bytes_with_project_info):
    info = _info(workbook_bytes_with_project_info)
    assert info["project_name"] == "Test Passive House"


def test_postal_code_strips_pphpp_quote_wrapping(workbook_bytes_with_project_info):
    """PHPP wraps zips in literal quotes ('"94028"') to keep leading zeros.
    The parser must strip them so the consumer gets a clean 5-digit string."""
    info = _info(workbook_bytes_with_project_info)
    assert info["postal_code"] == "94028"


def test_location_string_synthesized_from_address_cells(
    workbook_bytes_with_project_info,
):
    info = _info(workbook_bytes_with_project_info)
    assert info["location_string"] == "100 Main St, Sunnyvale, CA 94028"


def test_occupancy_type_via_phx_locator(workbook_bytes_with_project_info):
    """phi_building_use_type is read via PHX locator pattern. The raw
    PHX-formatted value (with `10-` prefix) is emitted as-is — the Mapper
    handles the human-friendly translation."""
    info = _info(workbook_bytes_with_project_info)
    assert info["occupancy_type"] == "10-Residential building: Residential"


def test_organizations_label_strips_role_suffix(workbook_bytes_with_project_info):
    """Overview labels look like 'Architect name / E-mail'. We keep the
    role prefix so the schema's `label` field carries something meaningful."""
    info = _info(workbook_bytes_with_project_info)
    by_label = {o["label"]: o["name"] for o in info["organizations"]}
    assert by_label["Architect"] == "Smith Architects"
    assert by_label["Mechanical engineer"] == "Cool MEP Inc"
    assert by_label["Energy consultant"] == "EnergyPro"
    assert by_label["Home owner"] == "Acme Owners LLC"
    assert by_label["Certification body"] == "PHIUS"


def test_organizations_skips_empty_rows(shape_en_10_6ip):
    """Real PHPP files have placeholder rows in the org block (e.g. unused
    Civil engineer slot). Empty-name rows must be dropped, not emitted as
    {label: '...', name: ''}."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ov_ws = wb.create_sheet(shape_en_10_6ip.OVERVIEW.name)
    # Only fill in row 33 (Architect)
    ov_ws["B33"] = "Architect name / E-mail"
    ov_ws["C33"] = "Solo Architect"

    buf = BytesIO()
    wb.save(buf)

    info = _info(buf.getvalue())
    assert info["organizations"] == [{"label": "Architect", "name": "Solo Architect"}]


def test_verification_complete_defaults_false(workbook_bytes_with_project_info):
    """P2.5 doesn't implement the verification_complete heuristic — emits
    False so the Mapper treats projects as design-stage by default. The
    follow-up ticket replaces this with a real Verification-sheet check."""
    info = _info(workbook_bytes_with_project_info)
    assert info["verification_complete"] is False


def test_handles_missing_overview_and_verification(workbook_bytes):
    """Workbook with neither Overview nor populated Verification — emit
    a fully-null project_info rather than crashing."""
    info = _info(workbook_bytes)
    assert info["project_name"] is None
    assert info["postal_code"] is None
    assert info["location_string"] is None
    assert info["occupancy_type"] is None
    assert info["verification_complete"] is False
    assert info["organizations"] == []


# --- v9.7 layout -----------------------------------------------------------
#
# v9.7 differs from v10 in both blocks this module reads, and in different
# ways: the Verification address block sits one row higher, and the Overview
# organization block sits at different rows AND uses different role labels.
# Hardcoded positions read a state as a zip and an interior-temperature
# setpoint as a firm name.


def _v97_verification():
    """v9.7's address block — one row higher than v10's."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Verification"
    ws["J4"] = "Building:"
    ws["K4"] = "Younis/Pang Residence"
    ws["J5"] = "Street:"
    ws["K5"] = "643 55th Street"
    ws["J6"] = "Postcode/City:"
    ws["K6"] = "95819"
    ws["L6"] = "Sacramento"
    ws["J7"] = "Province/Country:"
    ws["K7"] = "California"
    return wb


def test_address_row_found_by_label_not_position():
    ws = _v97_verification()["Verification"]
    assert _address_row(ws, "street") == 5
    assert _address_row(ws, "postcode/city") == 6
    assert _address_row(ws, "province/country") == 7


def test_v97_address_fields_are_not_shifted():
    """The regression: with v10's cells hardcoded, K7 is the postcode there
    and the province here, so a v9.7 file reported 'California' as its zip."""
    fields = _address_fields(_v97_verification()["Verification"])

    assert fields["street"] == "643 55th Street"
    assert fields["postal_code"] == "95819"
    assert fields["city"] == "Sacramento"
    assert fields["state_country"] == "California"


def test_address_fields_absent_labels_yield_none():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Verification"
    fields = _address_fields(wb["Verification"])
    assert all(v is None for v in fields.values())


# --- organization label vocabulary -----------------------------------------


@pytest.mark.parametrize(
    "label",
    [
        # v10.6
        "Home owner name / E-mail",
        "Architect name / E-mail",
        "Mechanical engineer name / E-mail",
        "Energy consultant name / E-mail",
        "Civil engineer name / E-mail",
        "Certification body Name / E-mail",
        # v9.7
        "Home owner / Client",
        "Architect",
        "Building services",
        "PHPP / Energy balance",
        "Building physics",
        "Structural engineering",
        "Certification body",
    ],
)
def test_real_role_labels_are_recognised(label):
    assert _is_organization_label(label)


@pytest.mark.parametrize(
    "label",
    [
        # These are what row constants actually picked up on v9.7 — the
        # reported "organizations" were 20 and 2.36, a temperature and a
        # heat-gain figure.
        "Interior temperatures winter / summer",
        "IHG winter / summer",
        "Specific values according to Passive House",
        "Treated floor area ATFA / Exterior volume",
        # near-misses that must not collide with "Building services" /
        # "Building physics"
        "Building type / Building status",
        "Building category, in terms of energy",
        "Building type / Construction",
        "Year of construction / Year of construction",
        "Type of certification",
    ],
)
def test_non_role_labels_are_rejected(label):
    assert not _is_organization_label(label)


def test_v97_organizations_read_names_not_setpoints(shape_en_10_6ip):
    """End-to-end over a v9.7-shaped Overview: the roles sit where v10 keeps
    unrelated numbers, so a positional reader returns '20' as a firm."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = shape_en_10_6ip.OVERVIEW.name
    rows = {
        26: ("Home owner / Client", "Maria Pang/Laith Younis"),
        27: ("Architect", "Bronwyn Barry/Passive House BB"),
        28: ("Building services", "Essential Air"),
        31: ("Structural engineering", None),  # empty name — skipped
        35: ("Interior temperatures winter / summer", "20"),
        36: ("IHG winter / summer", "2.36"),
    }
    for row, (label, name) in rows.items():
        ws[f"B{row}"] = label
        if name is not None:
            ws[f"C{row}"] = name

    orgs = _organizations(wb, shape_en_10_6ip)

    assert [o["name"] for o in orgs] == [
        "Maria Pang/Laith Younis",
        "Bronwyn Barry/Passive House BB",
        "Essential Air",
    ]
