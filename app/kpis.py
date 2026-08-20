"""KPI reads from a PHPP workbook (PAS-62 P2.2).

Emits the ``kpis`` subtree of the v1.0.0 output schema. Site EUI is
intentionally deferred — it requires summing PER T-column rows across
multiple end-use ranges, with high drift risk against Claude. Tracked
as a follow-up to P2.2.

Coverage in this module:
    tfa, heating_demand, cooling_demand,
    peak_loads.{heating,cooling},
    source_eui (PER total energy demand, V column),
    pe_demand   (PER total energy demand, X column)

Patterns used (validated by the 2026-04-24 spike):
    - direct (col, row) reads via the PHX shape (e.g. HEATING_DEMAND)
    - locator-string row scan (e.g. PER row 68 — "Total energy demand")

Sign convention: a negative cooling-peak or cooling-demand value is a
heat-loss number, not a real cooling demand. Return null in that case.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PHX.PHPP.phpp_localization import shape_model


def read_kpis(wb: Workbook, shape: shape_model.PhppShape) -> dict[str, Any]:
    """Read the kpis subtree.

    Always returns the full structure. A missing sheet or empty cell shows
    up as ``value: None`` on the affected measurement — we keep the unit
    and source so the consumer knows what the parser tried to read."""
    return {
        "tfa": _tfa(wb, shape),
        "heating_demand": _specific_demand(wb, shape.HEATING_DEMAND, "heating"),
        "cooling_demand": _specific_demand(wb, shape.COOLING_DEMAND, "cooling"),
        "peak_loads": {
            "heating": _peak_load(wb, shape.HEATING_PEAK_LOAD, sign="any"),
            "cooling": _peak_load(wb, shape.COOLING_PEAK_LOAD, sign="positive_only"),
        },
        "source_eui": _per_total(wb, shape.PER, column="per_energy"),
        "pe_demand": _per_total(wb, shape.PER, column="pe_energy"),
    }


def _tfa(wb: Workbook, shape: shape_model.PhppShape) -> dict[str, Any]:
    cd = shape.COOLING_DEMAND
    ws = _sheet(wb, cd.name)
    value = _num(ws[cd.address_tfa].value) if ws is not None else None
    return {
        "value": value,
        "unit": _area_unit(shape),
        "source": f"{cd.name}!{cd.address_tfa}",
    }


def _specific_demand(
    wb: Workbook,
    sub: shape_model.HeatingDemand | shape_model.CoolingDemand,
    kind: str,
) -> dict[str, Any]:
    row = (
        sub.row_annual_demand
        if kind == "heating"
        else sub.row_annual_sensible_demand
    )
    cell = f"{sub.col_kWh_m2_year}{row}"
    ws = _sheet(wb, sub.name)
    raw = _num(ws[cell].value) if ws is not None else None
    if kind == "cooling":
        raw = _positive_or_none(raw)

    return {
        "value": raw,
        "unit": _per_area_demand_unit(sub.unit),
        "source": f"{sub.name}!{cell}",
    }


def _peak_load(
    wb: Workbook,
    sub: shape_model.HeatingPeakLoad | shape_model.CoolingPeakLoad,
    sign: str,
) -> dict[str, Any]:
    """Peak load = max(weather_1, weather_2). PHPP computes both reference
    and project-specific weather; the design peak is the worse case."""
    if isinstance(sub, shape_model.HeatingPeakLoad):
        row = sub.row_total_load
    else:
        row = sub.row_total_sensible_load

    ws = _sheet(wb, sub.name)
    if ws is None:
        return {"value": None, "unit": _peak_load_unit(sub.unit)}

    w1 = _num(ws[f"{sub.col_weather_1}{row}"].value)
    w2 = _num(ws[f"{sub.col_weather_2}{row}"].value)
    value = _max_or_none(w1, w2)
    if sign == "positive_only":
        value = _positive_or_none(value)

    return {
        "value": value,
        "unit": _peak_load_unit(sub.unit),
        "source": f"{sub.name}!{sub.col_weather_1}{row}+{sub.col_weather_2}{row}",
    }


# v9.7 labels its summary columns in the column itself, and the strings differ
# from v10's ("PER demand" / "PE demand"). That difference is what makes the
# fallback in _per_total safe: these can only match a v9-shaped sheet.
_V9_SUMMARY_HEADERS = {
    "per_energy": "PER specific value",
    "pe_energy": "PE Value",
}

# On v9.7 the summary sits 3 rows under its header (units row, factor-set row,
# then the value) and the per-end-use breakdown starts at +5. The window stops
# short of the breakdown deliberately: if the summary is ever blank, this must
# return nothing rather than the heating row, which would be a plausible-looking
# fraction of the right answer.
_V9_SUMMARY_MAX_OFFSET = 4


def _per_total(
    wb: Workbook, per: shape_model.PER, column: str
) -> dict[str, Any]:
    """The PER grand total, which sits in a different place on v9 and v10.

    **v10** labels the row `Total energy demand <unit>` in the locator column
    and puts it at the BOTTOM, after the per-end-use breakdown. The unit suffix
    varies (kBTU/(ft²yr) vs kWh/(m²yr)) so we match by prefix. Hardcoding a row
    would break the moment a user adds heating types and the table shifts.

    **v9.7 has no such row at all** — the labels there are `Demand`,
    `Demand, cumulative generation (annual balance)` and so on, none of which
    start with "Total". Its summary sits at the TOP instead, between the header
    block and the breakdown, and is identified by a header in its own column
    rather than in the locator column. Before this was handled, source_eui and
    pe_demand came back null on every v9.7 file.

    The two layouts are inverted relative to the breakdown, so the v10 scan is
    tried first and the v9 fallback only runs when it finds nothing. A blind
    "first number under the header" rule would read v10's first BREAKDOWN row
    as its total.
    """
    ws = _sheet(wb, per.name)
    if ws is None:
        return {"value": None, "unit": _per_area_demand_unit(per.unit)}

    target_col = column_index_from_string(getattr(per.columns, column))

    row = _find_prefix_row(
        ws, column_index_from_string(per.locator_col), "Total energy demand"
    )
    if row is None:
        row = _find_v9_summary_row(ws, target_col, column)
    if row is None:
        return {"value": None, "unit": _per_area_demand_unit(per.unit)}

    cell_ref = f"{ws.cell(row=row, column=target_col).column_letter}{row}"
    return {
        "value": _num(ws.cell(row=row, column=target_col).value),
        "unit": _per_area_demand_unit(per.unit),
        "source": f"{per.name}!{cell_ref}",
    }


def _find_v9_summary_row(ws: Worksheet, target_col: int, column: str) -> int | None:
    """Row of the v9 PER summary for `column`, or None.

    Anchors on the header sitting in the target column itself, then takes the
    first numeric cell in a bounded window below it — the intervening rows hold
    the unit string and the factor-set dropdown, so no fixed offset is assumed.
    """
    header = _V9_SUMMARY_HEADERS.get(column)
    if header is None:
        return None

    header_row = _find_prefix_row(ws, target_col, header)
    if header_row is None:
        return None

    for offset in range(1, _V9_SUMMARY_MAX_OFFSET + 1):
        row = header_row + offset
        if _num(ws.cell(row=row, column=target_col).value) is not None:
            return row
    return None


# --- helpers ---------------------------------------------------------------


def _sheet(wb: Workbook, name: str) -> Worksheet | None:
    needle = name.strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == needle:
            return wb[sheet_name]
    return None


def _find_prefix_row(ws: Worksheet, col_idx: int, prefix: str) -> int | None:
    needle = prefix.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if isinstance(v, str) and v.strip().lower().startswith(needle):
            return r
    return None


def _num(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _positive_or_none(v: float | None) -> float | None:
    return v if isinstance(v, (int, float)) and v > 0 else None


def _max_or_none(a: float | None, b: float | None) -> float | None:
    candidates = [x for x in (a, b) if x is not None]
    return max(candidates) if candidates else None


def _area_unit(shape: shape_model.PhppShape) -> str:
    return "ft2" if shape.HEATING_DEMAND.unit.upper() == "KBTU" else "m2"


def _per_area_demand_unit(shape_unit: str) -> str:
    """Map PHX's unit string to the schema enum."""
    return "kBtu/ft2yr" if shape_unit.upper() == "KBTU" else "kWh/m2a"


def _peak_load_unit(shape_unit: str) -> str:
    u = shape_unit.upper().replace(" ", "")
    if u == "BTU/HR":
        return "Btu/h"
    if u == "KW":
        return "kW"
    return "W"
