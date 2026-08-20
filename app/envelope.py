"""Envelope reads from a PHPP workbook (PAS-62 P2.3).

Emits the ``envelope`` subtree of the v1.0.0 output schema. P2.3 covers:

  components  — walls / roofs / floors / doors from AREAS surface_rows,
                plus thermal bridges from AREAS thermal_bridge_rows
  airtightness.n50_ach — from VENTILATION (one-cell read)

Deferred to follow-up tickets:

  * P2.3.b — opaque U/R-values from R-Values sheet (assembly_id lookup
    via ``shape.UVALUES.constructor`` walk; non-trivial because each
    assembly is a multi-row block keyed by header offset).
  * P2.3.c — windows from the Windows sheet (separate sheet, separate
    conventions: u_w vs u_w_installed, glazing_id linkage).

Patterns:
  * iterate-until-blank from a locator header ("Area input", "Thermal
    bridge input"); termination when column L (description) is None.
  * group_number classification — column M holds e.g. "8-External wall -
    ambient"; we parse the integer prefix and map to the schema enum.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PHX.PHPP.phpp_localization import shape_model

# PHPP group_number → schema component enum. Pinned 2026-05-04 against the
# canonical 10.6 IP example. Groups 0 and 1 are aggregates (footprint, TFA);
# 2-6 are window placeholders that point at the Windows sheet (handled in
# P2.3.c). 7-11 are user-entered surface types.
GROUP_TO_COMPONENT = {
    7: "door",
    8: "wall",  # external wall - ambient
    9: "wall",  # ground wall (in contact with ground)
    10: "roof",
    11: "slab_on_grade",
}

_GROUP_PREFIX = re.compile(r"^\s*(\d+)\s*-")


def read_envelope(wb: Workbook, shape: shape_model.PhppShape) -> dict[str, Any]:
    """Read the envelope subtree. Always returns the structure; each
    sub-piece (components list, airtightness) handles its own missing-
    sheet case by emitting an empty/null variant."""
    return {
        "components": _components(wb, shape),
        "airtightness": _airtightness(wb, shape),
    }


def _components(wb: Workbook, shape: shape_model.PhppShape) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    out.extend(_surface_components(wb, shape.AREAS))
    out.extend(_thermal_bridge_components(wb, shape.AREAS))
    return out


def _surface_components(wb: Workbook, areas: shape_model.Areas) -> list[dict[str, Any]]:
    ws = _sheet(wb, areas.name)
    if ws is None:
        return []

    rows = areas.surface_rows
    header_row = _find_header_row(
        ws, rows.locator_col_header, rows.locator_string_header
    )
    if header_row is None:
        return []

    desc_col = column_index_from_string(rows.inputs.description.column)
    group_col = column_index_from_string(rows.inputs.group_number.column)

    # The shape's `area` input is the USER-OVERRIDE column ("User determined
    # [ft2]" on v9.7, "User-defined calculation [ft2]" on v10) -- empty for
    # every surface PHPP computes from its own a x b entry, which is most of
    # them. PHPP's result is a separate column headed "Area [ft2]", so prefer
    # that and keep the override as the fallback.
    area_col = _computed_area_col(ws, header_row + 1) or column_index_from_string(
        rows.inputs.area.column
    )
    area_letter = get_column_letter(area_col)

    out: list[dict[str, Any]] = []
    # Skip the header label row (header_row + 1 holds column labels like
    # "Building assembly description" — not data) and start at +2.
    for r in range(header_row + 2, ws.max_row + 1):
        desc = ws.cell(row=r, column=desc_col).value
        if desc is None or (isinstance(desc, str) and not desc.strip()):
            break  # iterate-until-blank termination

        group_raw = ws.cell(row=r, column=group_col).value
        component_type = _classify(group_raw)
        if component_type is None:
            continue  # group 0/1/2-6 — aggregates and window placeholders

        area = _num(ws.cell(row=r, column=area_col).value)
        out.append(
            {
                "component": component_type,
                "label": str(desc),
                "area_ft2": area,
                "source_area": f"{areas.name}!{area_letter}{r}",
            }
        )
    return out


def _thermal_bridge_components(
    wb: Workbook, areas: shape_model.Areas
) -> list[dict[str, Any]]:
    ws = _sheet(wb, areas.name)
    if ws is None:
        return []

    tbs = areas.thermal_bridge_rows
    header_row = _find_header_row(
        ws, tbs.locator_col_header, tbs.locator_string_header
    )
    if header_row is None:
        return []

    desc_col = column_index_from_string(tbs.inputs.description.column)
    length_col = column_index_from_string(tbs.inputs.length.column)
    psi_col = column_index_from_string(tbs.inputs.psi_value.column)

    out: list[dict[str, Any]] = []
    for r in range(header_row + 2, ws.max_row + 1):
        desc = ws.cell(row=r, column=desc_col).value
        if desc is None or (isinstance(desc, str) and not desc.strip()):
            break
        length = _num(ws.cell(row=r, column=length_col).value)
        psi = _num(ws.cell(row=r, column=psi_col).value)
        if length is None and psi is None:
            continue  # placeholder row with description only — skip
        out.append(
            {
                "component": "thermal_bridge",
                "label": str(desc),
                "length_ft": length,
                "psi_value_Btuh_ftF": psi,
                "source_area": f"{areas.name}!{tbs.inputs.length.column}{r}",
            }
        )
    return out


def _airtightness(
    wb: Workbook, shape: shape_model.PhppShape
) -> dict[str, Any]:
    """n50 lives at VENTILATION via the same locator-string pattern
    VERIFICATION uses for num_of_units (col I → input col M)."""
    vent = shape.VENTILATION
    n50 = vent.airtightness_n50

    ws = _sheet(wb, vent.name)
    if ws is None:
        return {"n50_ach": None, "source": None}

    locator_col = column_index_from_string(n50.locator_col)
    input_col = column_index_from_string(n50.input_column)
    needle = n50.locator_string.strip().lower()

    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=locator_col).value
        if isinstance(v, str) and v.strip().lower().startswith(needle):
            return {
                "n50_ach": _num(ws.cell(row=r, column=input_col).value),
                "source": f"{vent.name}!{n50.input_column}{r}",
            }

    return {"n50_ach": None, "source": None}


# --- helpers ---------------------------------------------------------------


def _classify(group_raw: Any) -> str | None:
    """Map an area-group cell to the schema component enum.

    Returns None for groups we intentionally skip (footprint/TFA/window
    placeholders) and for anything unrecognised, so the skip behaviour is the
    same whichever form the cell takes.
    """
    number = _group_number(group_raw)
    if number is None:
        return None
    return GROUP_TO_COMPONENT.get(number)


def _group_number(group_raw: Any) -> int | None:
    """The group number, from either form PHPP writes.

    v10 stores the labelled dropdown member as text -- ``"8-External wall -
    ambient"``. **v9.7 stores the bare number**, and openpyxl hands it back as
    an int. Accepting only the text form silently skipped every surface on
    every v9.7 file: the components list came back holding thermal bridges
    alone, which looks like a building with no walls rather than like a bug.
    """
    # bool is an int subclass, and a checkbox is not a group number.
    if isinstance(group_raw, bool):
        return None
    if isinstance(group_raw, int):
        return group_raw
    if isinstance(group_raw, float):
        return int(group_raw) if group_raw.is_integer() else None
    if isinstance(group_raw, str):
        match = _GROUP_PREFIX.match(group_raw)
        if match:
            return int(match.group(1))
        stripped = group_raw.strip()
        if stripped.isdigit():
            return int(stripped)
    return None


def _computed_area_col(ws: Worksheet, label_row: int) -> int | None:
    """Index of the column headed ``Area [<unit>]`` on the surface table's
    label row, or None.

    Located by label rather than by offset: it sits at AB on v9.7 and Z on
    v10, and the unit in the header varies with the workbook. The other
    area-bearing columns are all qualified ("User determined [ft2]",
    "Subtraction window areas [ft2]"), so the bare prefix is unambiguous --
    verified as a unique match on three real workbooks.
    """
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=label_row, column=col).value
        if isinstance(value, str) and " ".join(value.split()).lower().startswith(
            "area ["
        ):
            return col
    return None


def _find_header_row(ws: Worksheet, col_letter: str, needle: str) -> int | None:
    col_idx = column_index_from_string(col_letter)
    target = needle.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if isinstance(v, str) and v.strip().lower() == target:
            return r
    return None


def _sheet(wb: Workbook, name: str) -> Worksheet | None:
    needle = name.strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == needle:
            return wb[sheet_name]
    return None


def _num(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None
