"""Mechanical equipment, read from the workbook's defined names.

Fills the `hvac_equipment` subtree that schema/output.schema.json already
declares -- ParseResponse's docstring names this slot "P2.4 hvac". The schema
is strict (additionalProperties: false), so the key set in _item() is the contract,
not a convention.

Scope is PHPP v10.6. The names are stable across the IP and SI editions of a
version -- verified on two real workbooks -- but NOT across versions: v9.7
carries 91 WP_ ranges against v10.6's 43 with no overlap. Other versions
therefore emit nothing rather than guessing.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.workbook import Workbook

from app.named_ranges import resolve

# The v10.6 shapes this module understands. A version outside this set emits
# nothing -- see the module docstring on why names do not generalise.
SUPPORTED_VERSIONS = ("EN_10_6", "EN_10_6IP")

M3H_TO_CFM = 0.588578

# PHPP dropdown members are "<id>-<label>": "01ud-Swegon Casa R7",
# "1363vs03-Brink Climate Systems B.V. - Brink Flair". Anchored and
# non-greedy on the id so an internal hyphen in the label survives.
_PREFIX = re.compile(r"^[0-9]+[a-z]*[0-9]*-")


def _first(value: Any) -> Any:
    """A defined name may resolve to a scalar or a list; device selections are
    single-valued either way."""
    if isinstance(value, list):
        return value[0] if value else None
    return value


def _strip_prefix(text: Any) -> str | None:
    if not isinstance(text, str) or not text.strip():
        return None
    return _PREFIX.sub("", text.strip()) or None


def _classify_ventilation(humidity_recovery_pct: float | None) -> str:
    """A unit that recovers moisture is an ERV; heat-only is an HRV.

    Both units in the real corpus recover moisture, so this is the field that
    tells them apart -- and production currently stores one of them as `hrv`
    when its humidity recovery is 0.86.
    """
    if humidity_recovery_pct is None:
        return "hrv"
    return "erv" if humidity_recovery_pct > 0 else "hrv"


# The ventilation unit block on the Components sheet. Column letters are stable
# across the two editions; only the sheet name differs, and the named ranges
# already resolve to the metric one.
_VENT_SHEETS = ("Components SI", "Components")
_VENT_ID_COL = "LQ"
_VENT_HR_COL = "LS"
_VENT_HUMIDITY_COL = "LT"

# Two blocks, chosen by the device's id prefix. `NNud` is user-defined and sits
# in the user block; `NNNNvsNN` is a certified catalog entry. Searching only the
# catalog resolves certified units and silently misses every user-defined one.
_VENT_USER_ROWS = range(13, 114)
_VENT_CERT_ROWS = range(114, 915)


def _vent_sheet(wb: Workbook):
    for name in _VENT_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    return None


def _vent_spec(wb: Workbook, device_id: str | None) -> dict[str, Any]:
    """Heat and humidity recovery for a device, looked up by its id."""
    blank = {"heat_recovery_pct": None, "humidity_recovery_pct": None}
    ws = _vent_sheet(wb)
    if ws is None or not device_id:
        return blank

    rows = _VENT_CERT_ROWS if _is_certified_id(device_id) else _VENT_USER_ROWS
    for row in rows:
        try:
            if ws[f"{_VENT_ID_COL}{row}"].value != device_id:
                continue
            return {
                "heat_recovery_pct": _as_pct(ws[f"{_VENT_HR_COL}{row}"].value),
                "humidity_recovery_pct": _as_pct(ws[f"{_VENT_HUMIDITY_COL}{row}"].value),
            }
        except (KeyError, ValueError, IndexError):
            return blank
    return blank


def _is_certified_id(device_id: str) -> bool:
    """`1363vs03` is a catalog entry; `01ud` is user-defined."""
    return "ud" not in device_id


def _as_pct(value: Any) -> float | None:
    """PHPP stores these as a fraction (0.86); the schema wants a percentage."""
    if not isinstance(value, (int, float)):
        return None
    return float(value) * 100.0


def _device_id(raw: Any) -> str | None:
    """The id prefix of a dropdown member: '01ud-Swegon…' -> '01ud'."""
    if not isinstance(raw, str) or "-" not in raw:
        return None
    return raw.split("-", 1)[0].strip() or None


def _flag_set(wb: Workbook, name: str) -> bool:
    """PHPP marks an active device with an 'x' in an _Ankreuzen range."""
    value = _first(resolve(wb, name))
    return isinstance(value, str) and value.strip().lower() == "x"


def _item(**kwargs: Any) -> dict[str, Any]:
    """Every emitted item carries the full key set, nil included.

    The schema forbids extra keys, and a consumer that assign_attributes-es a
    partial hash leaves whatever the previous extraction wrote at that index --
    the same reasoning Mechanical::EquipmentAttributes documents on the Ruby
    side.
    """
    base: dict[str, Any] = {
        "equipment_type": None, "name": None, "manufacturer": None,
        "capacity": None, "capacity_unit": None,
        "efficiency_value": None, "efficiency_type": None,
        "airflow_cfm": None, "airflow_m3h": None,
        "heat_recovery_efficiency_pct": None, "source": None,
    }
    base.update(kwargs)
    return base


def read_equipment(wb: Workbook, version: str) -> list[dict[str, Any]]:
    """Every active mechanical device, or [] for an unsupported version."""
    if version not in SUPPORTED_VERSIONS:
        return []

    items: list[dict[str, Any]] = []
    items.extend(_ventilation(wb))
    items.extend(_cooling(wb))
    items.extend(_heat_pumps(wb))
    return items


def _ventilation(wb: Workbook) -> list[dict[str, Any]]:
    source = "Lueftung_Auswahl_Lueftungsgeraet"
    raw = _first(resolve(wb, source))
    name = _strip_prefix(raw)
    if name is None:
        return []

    spec = _vent_spec(wb, _device_id(raw))
    airflow_m3h = _as_float(_first(resolve(wb, "Lueftung_Auslegungsvolumenstrom")))

    # Manufacturer is always None; see comment in _cooling -- that ruling holds
    # here too, so it is deliberately absent from this call, not forgotten.
    return [_item(
        equipment_type=_classify_ventilation(spec["humidity_recovery_pct"]),
        name=name,
        heat_recovery_efficiency_pct=spec["heat_recovery_pct"],
        airflow_m3h=airflow_m3h,
        airflow_cfm=None if airflow_m3h is None else airflow_m3h * M3H_TO_CFM,
        source=source,
    )]


def _as_float(value: Any) -> float | None:
    return float(value) if isinstance(value, (int, float)) else None


def _cooling(wb: Workbook) -> list[dict[str, Any]]:
    if not _flag_set(wb, "Kuehlgeraete_Umluft_Kuehlung_Ankreuzen"):
        return []
    source = "Kuehlgeraete_Kompressor_Umluft_Geraet"
    name = _strip_prefix(_first(resolve(wb, source)))
    if name is None:
        return []
    # Manufacturer is always None: any first-token heuristic invents sometimes,
    # and the available corpus shows half the devices select PHPP generics
    # (e.g., "Standard air-to-air heat pump"). A wrong manufacturer on a
    # certifier-facing record is worse than a blank one; the full name still
    # carries identifying info.
    return [_item(
        equipment_type="cooling_unit",
        name=name,
        source=source,
    )]


def _heat_pumps(wb: Workbook) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for source, equipment_type in (
        ("WP_Heizungssystem_Auswahl_Luft_Luft_WP", "heat_pump"),
        ("WP_Warmwassersystem_Auswahl_WP", "dhw_heat_pump"),
    ):
        name = _strip_prefix(_first(resolve(wb, source)))
        if name is None:
            continue
        out.append(_item(
            equipment_type=equipment_type,
            name=name,
            source=source,
        ))
    return out
