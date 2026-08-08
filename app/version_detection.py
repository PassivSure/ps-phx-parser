"""Detect a PHPP workbook's version + language from its Data sheet.

Ports PHX.PHPP.phpp_app.PHPPConnection.get_phpp_version to headless
openpyxl so we can run on Heroku without Excel/xlwings. See
spikes/version-detection.md for the writeup + empirical validation.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DATA_SHEET_NAMES = ("DATA", "DATEN", "DATOS")

# easyPH workbooks append an edition tag to the version cell on 'Data':
#
#     standard PHPP   "10.6"           /  "10.6 IP"
#     easyPH          "10.6 easyPHv3"  /  "10.6 easyPHv3 IP"
#
# The tag names the EDITION, not a different shapefile. An easyPH file is a
# standard PHPP plus an 'easyPH' input worksheet — same sheets, same cell
# geometry, same localization — so it uses its base version's shape. PHX itself
# treats it that way: PHPPConnection.is_easyPh() detects the edition by looking
# for the worksheet, never by reading this string.
#
# Left in, the tag lands in `minor` and shape_stem yields EN_10_6EASYPHV3IP —
# a shapefile that does not and should not exist.
EASYPH_EDITION_TAG = re.compile(r"\s*easyPH\s*v?\d*(?:\.\d+)*\s*", re.IGNORECASE)
LANGUAGE_PROXIES = {
    "1-PE-FAKTOREN": "DE",
    "1-FACTORES EP": "ES",
    "1-PE-FACTORS": "EN",
}


class DetectionError(Exception):
    """Raised when the workbook's version can't be determined."""


@dataclass(frozen=True)
class DetectedVersion:
    major: str
    minor: str
    language: str
    raw: str  # the raw "10.6 IP" string read from the workbook

    @property
    def shape_stem(self) -> str:
        """PHX's `{LANG}_{major}_{minor}` filename stem.

        Matches PHX.PHPP.phpp_model.version.PHPPVersion.clean_input:
        upper-case, strip, remove spaces, replace dots with underscores.
        """
        major = _clean(self.major)
        minor = _clean(self.minor)
        return f"{self.language}_{major}_{minor}"


def _clean(value: str) -> str:
    return value.upper().strip().replace(" ", "").replace(".", "_")


def strip_easyph_edition_tag(raw_version: str) -> str:
    """Drop any 'easyPH' edition tag, leaving the rest of the string intact.

    The unit-system suffix survives, because that IS part of the shape:

        "10.6 easyPHv3"     -> "10.6"      -> EN_10_6
        "10.6 easyPHv3 IP"  -> "10.6 IP"   -> EN_10_6IP
        "10.6 IP"           -> "10.6 IP"   (unchanged)
    """
    return EASYPH_EDITION_TAG.sub(" ", str(raw_version)).strip()


def _find_data_sheet(wb) -> Worksheet:
    sheets_upper = {s.upper(): s for s in wb.sheetnames}
    for candidate in DATA_SHEET_NAMES:
        if candidate in sheets_upper:
            return wb[sheets_upper[candidate]]
    raise DetectionError(
        f"No Data/Daten/Datos sheet. Sheets: {wb.sheetnames[:8]}"
    )


def _find_phpp_row(ws: Worksheet, col: int = 1, max_row: int = 10) -> int:
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and str(v).upper().strip().replace(" ", "").startswith("PHPP"):
            return r
    raise DetectionError(
        f"No 'PHPP' marker in col A rows 1-{max_row} of sheet {ws.title!r}. "
        "Likely a pre-v9 PHPP — not supported."
    )


def detect_version(workbook_bytes: bytes) -> DetectedVersion:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    try:
        ws = _find_data_sheet(wb)
        row = _find_phpp_row(ws)
        row_vals = [ws.cell(row=row, column=c).value for c in range(1, 20)]
        non_blank = [v for v in row_vals if v not in (None, "")]

        if len(non_blank) < 2:
            raise DetectionError(f"PHPP row {row} has no version cell: {non_blank}")

        raw_version = str(non_blank[1])
        # easyPH tags the edition onto the version; it shares its base version's
        # shapefile, so drop the tag before splitting. `raw` keeps the original
        # so callers can still see the workbook was an easyPH edition.
        parsed_version = strip_easyph_edition_tag(raw_version)
        if "." not in parsed_version:
            raise DetectionError(f"Unexpected version cell {raw_version!r}")
        major, minor = parsed_version.split(".", 1)

        language = _detect_language(non_blank)

        return DetectedVersion(
            major=major, minor=minor, language=language, raw=raw_version
        )
    finally:
        wb.close()


def _detect_language(row_non_blank: list) -> str:
    """Language defaults to EN.

    PHPP v10+ Data row has a PE-factor proxy as the last non-blank cell,
    which we match against LANGUAGE_PROXIES. Pre-v10 files lack it — since
    we only ship EN shape files, defaulting to EN is the pragmatic fallback.
    """
    last = str(row_non_blank[-1]).upper().strip()
    for needle, lang in LANGUAGE_PROXIES.items():
        if needle in last:
            return lang
    return "EN"
