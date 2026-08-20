"""Project info reads from a PHPP workbook (PAS-62 P2.5).

Emits the ``project_info`` subtree of the v1.0.0 output schema. Mixes two
sources because PHX's shape doesn't map all the fields ps-rails consumes:

  PHX-mapped (locator pattern):
    project_name     — OVERVIEW.basic_data.address_project_name (Overview!C11)
    occupancy_type   — VERIFICATION.phi_building_use_type
                        (locator R="Building use", value T at row+1)

  Direct cell reads (PHX gap):
    postal_code      — Verification!K7 (strip surrounding quotes)
    location_string  — synthesized from Verification!K6/L7/K8/K7
    organizations    — Overview rows 31-36 + 48: column B has the role label,
                        column C has the firm/person name

  Deferred:
    verification_complete — needs a heuristic over Verification's certifier
                              fields. Filed as a follow-up; emits False here.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PHX.PHPP.phpp_localization import shape_model

ORGANIZATION_LABEL_COL = "B"
ORGANIZATION_NAME_COL = "C"

# Role labels in Overview column B, as a prefix each row must start with.
# The two versions name the same block differently and put it in different
# places -- v10 has "Architect name / E-mail" at row 33, v9.7 has "Architect"
# at row 27 -- so the union of both vocabularies is scanned and the row is
# found by its label. Row constants pinned against one file read v9.7's
# "Interior temperatures winter / summer" as an organization and reported its
# value, 20, as the firm's name.
#
# Prefixes must stay specific enough not to collide: "Building services" and
# "Building physics" are roles, while "Building type / Building status" and
# "Building category" are not, and none of the four is a prefix of another.
ORGANIZATION_LABEL_PREFIXES = (
    # shared
    "home owner",
    "architect",
    "certification body",
    # v10.6
    "mechanical engineer",
    "energy consultant",
    "civil engineer",
    # v9.7
    "building services",
    "phpp / energy balance",
    "building physics",
    "structural engineering",
    "contractor / tradesperson",
)

# Row bound for the Overview scan. The block sits in the twenties/forties on
# both versions; stopping well before the KPI block below it keeps a stray
# match from reaching numbers.
ORGANIZATION_SCAN_MAX_ROW = 60

# Address fields live in the Verification header, which PHX's shape does not
# map. They are found by the label in column J rather than by cell, because
# **the v9.7 block sits exactly one row higher than v10's** -- the same 1-row
# shift the verification results block carries. Hardcoded to v10, "K7" is the
# postcode there and the province here, so a v9.7 file reported its state as
# its zip code.
ADDRESS_LABEL_COL = "J"
ADDRESS_LABELS = {
    "street": "street",
    "postcode_city": "postcode/city",
    "state_country": "province/country",
}
ADDRESS_SCAN_MAX_ROW = 30


def read_project_info(wb: Workbook, shape: shape_model.PhppShape) -> dict[str, Any]:
    return {
        "project_name": _project_name(wb, shape),
        "postal_code": _postal_code(wb, shape),
        "location_string": _location_string(wb, shape),
        "occupancy_type": _occupancy_type(wb, shape),
        "verification_complete": False,  # deferred — see module docstring
        "organizations": _organizations(wb, shape),
    }


def _project_name(wb: Workbook, shape: shape_model.PhppShape) -> str | None:
    ws = _sheet(wb, shape.OVERVIEW.name)
    if ws is None:
        return None
    address = shape.OVERVIEW.basic_data.address_project_name
    return _str(ws[address].value)


def _address_row(ws: Worksheet, label: str) -> int | None:
    """Row whose column-J label starts with `label`, or None."""
    col = column_index_from_string(ADDRESS_LABEL_COL)
    needle = label.strip().lower()
    for row in range(1, min(ws.max_row, ADDRESS_SCAN_MAX_ROW) + 1):
        value = ws.cell(row=row, column=col).value
        if isinstance(value, str) and value.strip().lower().startswith(needle):
            return row
    return None


def _address_fields(ws: Worksheet) -> dict[str, str | None]:
    """Street / postcode / city / state, located by their own labels.

    PHPP puts the postcode and city on one row (K and L) under a single
    "Postcode/City:" label, and the province in K of the next labelled row.
    """
    fields: dict[str, str | None] = {
        "street": None,
        "postal_code": None,
        "city": None,
        "state_country": None,
    }

    street_row = _address_row(ws, ADDRESS_LABELS["street"])
    if street_row:
        fields["street"] = _str(ws.cell(row=street_row, column=11).value)  # K

    pc_row = _address_row(ws, ADDRESS_LABELS["postcode_city"])
    if pc_row:
        fields["postal_code"] = _str(ws.cell(row=pc_row, column=11).value)  # K
        fields["city"] = _str(ws.cell(row=pc_row, column=12).value)  # L

    state_row = _address_row(ws, ADDRESS_LABELS["state_country"])
    if state_row:
        fields["state_country"] = _str(ws.cell(row=state_row, column=11).value)

    return fields


def _postal_code(wb: Workbook, shape: shape_model.PhppShape) -> str | None:
    ws = _sheet(wb, shape.VERIFICATION.name)
    if ws is None:
        return None
    raw = _address_fields(ws)["postal_code"]
    if raw is None:
        return None
    # PHPP wraps zips in literal quotes (e.g. '"03049"') to keep leading
    # zeros from being parsed as numbers.
    return raw.strip().strip('"').strip("'") or None


def _location_string(wb: Workbook, shape: shape_model.PhppShape) -> str | None:
    """Synthesize "<street>, <city>, <state> <zip>" from the Verification
    header. Returns None if street + city + state are all missing."""
    ws = _sheet(wb, shape.VERIFICATION.name)
    if ws is None:
        return None

    fields = _address_fields(ws)
    street = fields["street"]
    city = fields["city"]
    state = fields["state_country"]
    zip_ = _postal_code(wb, shape)

    parts: list[str] = []
    if street:
        parts.append(street)
    if city and state:
        parts.append(f"{city}, {state}{f' {zip_}' if zip_ else ''}")
    elif city:
        parts.append(city)
    elif state:
        parts.append(state)

    return ", ".join(parts) if parts else None


def _occupancy_type(wb: Workbook, shape: shape_model.PhppShape) -> str | None:
    """Look up phi_building_use_type via the PHX locator pattern. Returns
    the raw PHX-formatted string (e.g. ``"21-Non-res building: School ..."``);
    the Mapper does the human-friendly translation against OccupancyType."""
    ws = _sheet(wb, shape.VERIFICATION.name)
    if ws is None:
        return None

    item = shape.VERIFICATION.phi_building_use_type
    locator_col = column_index_from_string(item.locator_col)
    needle = item.locator_string.strip().lower()

    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=locator_col).value
        if isinstance(v, str) and v.strip().lower() == needle:
            target_row = r + item.input_row_offset
            target_col = column_index_from_string(item.input_column)
            return _str(ws.cell(row=target_row, column=target_col).value)
    return None


def _organizations(
    wb: Workbook, shape: shape_model.PhppShape
) -> list[dict[str, str]]:
    """Overview carries a role label in column B and the firm/person name in
    column C. Rows are found by their label, not by position -- see
    ORGANIZATION_LABEL_PREFIXES. Empty-name rows are skipped."""
    ws = _sheet(wb, shape.OVERVIEW.name)
    if ws is None:
        return []

    out: list[dict[str, str]] = []
    label_col = column_index_from_string(ORGANIZATION_LABEL_COL)
    name_col = column_index_from_string(ORGANIZATION_NAME_COL)
    for row in range(1, min(ws.max_row, ORGANIZATION_SCAN_MAX_ROW) + 1):
        label = _str(ws.cell(row=row, column=label_col).value)
        if not label or not _is_organization_label(label):
            continue
        name = _str(ws.cell(row=row, column=name_col).value)
        if not name:
            continue
        out.append({"label": _clean_org_label(label) or "", "name": name})
    return out


def _is_organization_label(label: str) -> bool:
    normalized = " ".join(label.split()).strip().lower()
    return any(normalized.startswith(p) for p in ORGANIZATION_LABEL_PREFIXES)


_LABEL_TRAIL = re.compile(r"\s*[/]\s*.*$")  # strip "/ E-mail" / "/ Last name"


def _clean_org_label(label: str | None) -> str | None:
    """Overview labels are formatted ``"Architect name / E-mail"``. Keep
    the role prefix (everything before the slash) so the schema's `label`
    field carries something meaningful."""
    if not label:
        return None
    # Drop "/ <secondary>" tail
    role = _LABEL_TRAIL.sub("", label).strip()
    # Drop trailing " name"
    return re.sub(r"\s+name\s*$", "", role, flags=re.IGNORECASE) or None


# --- helpers ---------------------------------------------------------------


def _sheet(wb: Workbook, name: str) -> Worksheet | None:
    needle = name.strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == needle:
            return wb[sheet_name]
    return None


def _str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None
