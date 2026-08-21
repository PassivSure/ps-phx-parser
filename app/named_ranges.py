"""Resolve a workbook's defined names to their values.

PHPP addresses most of its equipment data by defined name rather than by cell.
Defined names in PHPP are authored to point at the METRIC pane: on an IP workbook
the unprefixed names point at `Ventilation SI`, `HP SI` and so on; on an SI
workbook they point at the unsuffixed sheets holding the same data. Downstream
callers read equipment from those named ranges and expect metric values.

This function retrieves whatever value the name points to, without imposing
unit guarantees -- that is the caller's responsibility. Equipment reads go
through here rather than through a per-edition cell map because the named
references themselves abstract the pane choice, whereas a cell map would need
one for each edition. That is why choosing the wrong pane was the defect class
that produced four separate fixes in the EN_9_7IP work.
"""

from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)

# "'Ventilation SI'!$K$91:$M$91" or "Ventilation!$B$4"
_REF = re.compile(r"^'?([^'!]+)'?!(\$?[A-Z]+\$?\d+)(?::(\$?[A-Z]+\$?\d+))?$")


def resolve(wb: Workbook, name: str) -> Any | None:
    """The value(s) behind a defined name, or None.

    The return shape is keyed on the reference's syntax, not on which cells are
    populated:
    - A reference WITHOUT `:` (e.g., 'Sheet'!$B$4) returns a scalar.
    - A reference WITH `:` (e.g., 'Sheet'!$K$91:$M$91) returns a list of
      non-empty values, or None if every cell is empty. This includes degenerate
      ranges like 'Sheet'!$B$4:$B$4 (same cell on both sides).

    This distinction matters: a caller writing `float(resolve(wb, name))` will work
    where a range has 2+ populated cells and raise TypeError where it has exactly
    one. Callers must therefore normalize via a `_first()` helper rather than
    assume they know the shape in advance.

    Roughly half of PHPP's equipment names are ranges, and a resolver that silently
    took the first cell would be right often enough to pass a casual test and wrong
    wherever the value sits elsewhere in the range -- hence the distinction.

    Returns None rather than raising for most failure modes -- an absent name,
    a sheet the workbook does not have, an unparseable reference such as #REF!.
    NOT caught here: a reference whose row exceeds openpyxl's max of 1048576
    raises ValueError out of the underlying cell access. Equipment is one
    subtree of five and must never take a parse down with it, but that guard
    lives in app.equipment.read_equipment (this function's only caller)
    rather than here, deliberately.
    """
    defined = wb.defined_names.get(name)
    if defined is None:
        return None

    match = _REF.match(str(defined.value).strip())
    if match is None:
        return None

    sheet_name, first, last = match.groups()
    if sheet_name not in wb.sheetnames:
        # An absent NAME is ordinary -- most of PHPP's names are unused in any
        # given file, and logging those would be pure noise. A name that exists
        # and points at a sheet the workbook does not have is not ordinary: it
        # means the mapping and the file disagree, which is worth seeing.
        logger.warning(
            "[equipment] defined name %r points at missing sheet %r", name, sheet_name
        )
        return None
    ws = wb[sheet_name]

    if last is None:
        return ws[first].value

    values = [
        cell.value
        for row in ws[f"{first}:{last}"]
        for cell in row
        if cell.value is not None
    ]
    return values or None
